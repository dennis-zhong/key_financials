"""
Pull Key Financials from Cap IQ screenshot to formatted Excel Sheet
"""

def makeKeyFinancials(lstofdata, ticker):
    from shutil import copyfile
    from copy import copy
    import openpyxl

    # ticker = str(input("Input a ticker:\n"))
    # while isBadTicker(ticker):
    #     print("Invalid input")
    print("Making Key_Financials...")
    copyfile("./templates/Key_Financials_Template.xlsx", "./sheets/Key_Financials_"+ticker+".xlsx")
    wb = openpyxl.load_workbook("./sheets/Key_Financials_"+ticker+".xlsx")
    ws = wb.active
    ws["A1"] = ticker +" Key Financials In Millions of USD, except per share items."
    i = 2 # starting index for row
    while ws["A"+str(i)].value == None: # needs to be blank row with 12 months following it
        i+=1
    if(lstofdata[0][0]==""):
        i-=1
    while lstofdata:
        currrow = lstofdata.pop(0)
        ws["A"+str(i)] = currrow.pop(0)
        start = ""
        columns = list("BCDEFGHIJKLMNOPQRSTUVWXYZ")
        for x in columns:
            if ws[x+str(i)].value != None:
                start = x
                break
        columns = columns[columns.index(start):]
        while currrow:
            currcol = columns.pop(0)
            curritem = currrow.pop(0)
            ws[currcol+str(i)] = curritem
        i+=1
        while ws["A"+str(i)].value == None and lstofdata:
            i+=1
    listofcolumns = list("JKLMNOPQRSTUVWXYZ")[::-1] # trying to find columns that need formatting
    for x in listofcolumns:
        if(ws[x+str(i-1)].value):
            break
        listofcolumns = listofcolumns[1:]
    for x in listofcolumns: # formatting columns
        for row in range(1, i):
            ws[x+str(row)]._style = copy(ws["A"+str(row)]._style)
    wb.save("./sheets/Key_Financials_"+ticker+".xlsx")
    print("Finished!\n")

def pytess():
    import pytesseract
    import cv2
    import os

    pngs = os.listdir("pics")
    for png in pngs:
        print(f"Reading {png}...")
        img = cv2.imread(f"pics/{png}")
        unformattedstr = pytesseract.image_to_string(img)
        lstofdata = unformattedstr.split("\n")[:-1]
        categories = [[""]]
        index = 0
        for x in lstofdata:
            if not x:
                continue # empty strings
            index+=1
            if x[0].isdigit():
                break
            categories.append([x])
        lstofdata = [x for x in lstofdata[index:] if x]
        for i, x in enumerate(lstofdata):
            curr = i%len(categories)
            categories[curr].append(x)
        makeKeyFinancials(categories, png[:png.index(".")])

from constants import *
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time

def check_alive(driver):
    try:
        if driver.service.process: # checks if process is alive
            return False
        else:
            return True
    except:
        return False # if errors out

def enter_credentials(driver):
    driver.find_element_by_id("password").send_keys(PASSWORD)
    driver.find_element_by_id("username").send_keys(USERNAME)
    driver.find_element_by_name("_eventId_proceed").click() # sign in

def verify_input():
    x = input("Enter ticker: ")
    count = 0
    while(count<3):
        print("Verify these tickers: (Y/N) " + x)
        verify = input()
        count+=1
        if(verify=="Y" or verify=="y"):
            return x
        elif(verify=="N" or verify=="n"):
            x = input("Enter ticker: ")
    print("Exiting program")
    return ""


def access_page(driver, ticker):
    # if(driver.title != "Welcome Georgetown University > Dashboard"):
    #     driver.refresh()
    driver.find_element_by_id("SearchTopBar").send_keys(ticker)
    driver.find_element_by_id("ciqSearchSearchButton").click()
    if(driver.title=="Search Profiles"):
        print("Unable to find proper ticker. Please choose from below list.")
        lst_of_poss = []
        for x in range(3):
            curr_elem = driver.find_element_by_id("SR"+str(x))
            lst_of_poss.append(curr_elem)
            print(str(x)+": "+curr_elem.text+"\n")
        # can add a "more" option
        pick = input()
        count = 0
        while(count<3):
            print("Verify your pick: (Y/N) " + pick)
            verify = input()
            count+=1
            if(verify=="Y" or verify=="y"):
                break
            elif(count==3):
                print("Exiting program")
            elif(verify=="N" or verify=="n"):
                pick = input("Enter your pick: ")
        lst_of_poss[int(pick)].find_element_by_tag_name("a").click()

def process_tables(driver):
    table = driver.find_element_by_id("_keyFinSection_ctl00_0_gv_0").text.split('\n') # Key Stats table
    table = table[2:table.index("            Currency")] # get rid of unnecessary text and space
    header = table[:table.index("Currency")] # get the top line so can format
    currency_index = max(loc for loc, val in enumerate(table) if val == "USD")+1

    textlst = table[table.index("Currency"):currency_index] # gets currency line

    table = table[currency_index:]
    tablelst = []
    currlst = []
    for i, x in enumerate(table):
        if(i%2==0):
            currlst.append(x)
        else:
            currlst+=x.split()
            tablelst+=[currlst]
            currlst = []
    headerlst = []
    headerlst.append(header.pop(0)) # process header
    line = ""
    while header:
        curr = header.pop(0)
        if("-" in curr):
            headerlst.append(line+curr)
            line = ""
        else:
            line+=curr+"\n"
    return [headerlst]+[textlst]+tablelst

def to_desired_page(driver, depth):
    try:
        driver.find_element_by_id("ll_7_123_2083").click()
        return 0
    except Exception as e:
        print(e)
        if(depth>=3):
            print("Unable to resolve to_desired_page")
            return 1
        curr_time = time.time()
        if(time.time() - curr_time)>3:
            print("Retrying key stats")
            to_desired_page(driver, depth+1)


stages = {
    "stage 1": lambda driver: driver.get("http://proxy.library.georgetown.edu/login?url=http://na.capitaliq.com/ip/GTOWU"),
    "stage 2": enter_credentials,
    "stage 3": verify_input,
    "stage 4": access_page,
    "stage 5": to_desired_page, # Key Stats button
    "stage 6": process_tables
}

def auto_scrape():
    try:
        with open("chrome_instances.txt") as chromes:
            url = chromes.readline()
            session_id = chromes.readline()
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        chrome_options.add_argument("--incognito")
        driver = webdriver.Chrome(executable_path=CHROME_PATH, options=chrome_options)
        if(session_id and url):
            save_url = driver.command_executor._url # in case chrome_instances don't work
            save_session = driver.session_id
            try:
                driver.close()
                driver.command_executor._url = url
                driver.session_id = session_id # restore
                driver.title # test out new driver
            except Exception as e:
                print("Past Driver unable to reuse")
                print(e)
                message = e.__str__()
                driver.command_executor._url = save_url
                driver.session_id = save_session
                driver.start_session({})
        if("capital-iq" not in driver.current_url):
            stages["stage 1"](driver)
        if(driver.title=="Single Signon - Georgetown University"):
            stages["stage 2"](driver)
        current_time = time.time()
        times_printed = 0
        while(driver.title=="Single Signon - Georgetown University"):
            if(int(time.time()-current_time)//3>times_printed):
                print("Waiting for DUO...")
                times_printed+=1
            elif(time.time()-current_time>20):
                print("Duo unresponsive. Exiting program.")
                return 1
        tickers = ""
        if(driver.title!="Single Signon - Georgetown University"):
            tickers = stages["stage 3"]()
            tickers = tickers.split(" ")
        for ticker in tickers:
            stages["stage 4"](driver, ticker) # go to ticker page
            if stages["stage 5"](driver, 0): # go to key stats
                continue
            processed_table = stages["stage 6"](driver) # process table
            makeKeyFinancials(processed_table, ticker)
    except Exception as e:
        print("Error with auto_scrape: ")
        print(e)
    if(driver.service.process or "Connection refused" in message):
        with open("chrome_instances.txt", "w") as chromes:
            chromes.write(driver.command_executor._url+"\n")
            chromes.write(driver.session_id)
    else:
        driver.stop_client()


def main():
    # pytess() this is for screenshots in pics
    auto_scrape()

if __name__ == "__main__":
    main()
